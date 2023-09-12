import xlrd
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
workbook1=xlrd.open_workbook(r'C:\Users\Administrator\Desktop\tablet\甲状腺组织学2018-2023.xls')  #打开表格1(手术及蜡块混合表)
workbook2=xlrd.open_workbook(r'C:\Users\Administrator\Desktop\tablet\细胞学2018-2023.xls')  #打开表格2（细胞液基涂片表）
sheet1=workbook1.sheet_by_index(0)    #读取表格1的第一个表
sheet2=workbook2.sheet_by_index(0)    #读取表格2的第一个表
a1=1
b1=sheet1.nrows   #表格1的第一个表的总行数#
b1
c1=[]
lk='蜡块'
while a1<b1:
    xxx1=sheet1.cell_value(rowx=a1,colx=22)
    if lk in xxx1:
        c1.append(a1)
    a1+=1
c1
workbook=xlsxwriter.Workbook('C:\\Users\\Administrator\\Desktop\\tablet\\waxblock.xlsx')
worksheet=workbook.add_worksheet('Sheet1')
#添加蜡块信息到新的表中#
x0=sheet1.row_values(0)
i0=0
while i0<len(x0):
    worksheet.write(0,i0,x0[i0])
    i0+=1

c1len=len(c1)
i=1
while i<c1len:
    ii=0
    xx=sheet1.row_values(c1[i])
    xxlen=len(xx)
    while ii<xxlen:
        worksheet.write(i,ii,xx[ii])
        ii+=1
    i+=1
workbook.close()
cx=list(range(1,b1-1))
c2=[i for i in cx if i not in c1]
workbook=xlsxwriter.Workbook('C:\\Users\\Administrator\\Desktop\\tablet\\ops.xlsx') #创建新的表格保存细胞蜡块信息
worksheet=workbook.add_worksheet('Sheet1')                                              #添加手术信息到新的表中
x0=sheet1.row_values(0)
i0=0
while i0<len(x0):
    worksheet.write(0,i0,x0[i0])
    i0+=1

c2len=len(c2)
i=1
while i<c2len:
    ii=0
    xx=sheet1.row_values(c2[i])
    xxlen=len(xx)
    while ii<xxlen:
        worksheet.write(i,ii,xx[ii])
        ii+=1
    i+=1
    i
workbook.close()
workbookops=openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\tablet\ops.xlsx')
workbookwax=openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\tablet\waxblock.xlsx')
sheetops=workbookops.active
sheetwax=workbookwax.active
i=2
opslen=sheetops.max_row +1
waxlen=sheetwax.max_row +1
opslen
waxlen
colx=sheetops.max_column
colx
#对比蜡块,添加细胞涂片信息#
while i<opslen :
    kk=sheetops.cell(i,8).value
    ii=1
    while ii<waxlen:
        kx=sheetwax.cell(ii,8).value
        if kx==kk:
            sheetops.cell(i,colx+1).value=sheetwax.cell(ii,4).value
            sheetops.cell(i,colx+2).value=sheetwax.cell(ii,23).value
            sheetops.cell(i,colx+3).value=sheetwax.cell(ii,14).value
        ii+=1
    i+=1       
workbookops.save(r'C:\Users\Administrator\Desktop\tablet\ops1.xlsx')
workbookops.close()
workbookops=openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\tablet\ops1.xlsx')
workbookwax=xlrd.open_workbook(r'C:\Users\Administrator\Desktop\tablet\细胞学2018-2023.xls')
sheetops=workbookops.active
sheetwax=workbookwax.sheet_by_index(0)
i=2
opslen=sheetops.max_row +1
waxlen=sheetwax.nrows
opslen
waxlen
colx=sheetops.max_column
colx
#对比细胞蜡块记录,添加细胞蜡块信息#
while i<opslen :
    kk=sheetops.cell(i,8).value
    nn=sheetops.cell(i,colx).value
    ii=1
    while ii<waxlen:
        kx=sheetwax.cell_value(rowx=(ii-1),colx=7)
        nx=sheetwax.cell_value(rowx=(ii-1),colx=8)
        if kx==kk and nx==nn:
            sheetops.cell(i,colx+1).value=sheetwax.cell_value(rowx=(ii-1),colx=3)
            sheetops.cell(i,colx+2).value=sheetwax.cell_value(rowx=(ii-1),colx=12)
            sheetops.cell(i,colx+3).value=sheetwax.cell_value(rowx=(ii-1),colx=8)
            sheetops.cell(i,colx+4).value=sheetwax.cell_value(rowx=(ii-1),colx=11)
        ii+=1
    i+=1
sheetops.cell(1,colx+4).value='细胞送检时间'
sheetops.cell(1,colx+3).value='细胞送检部位'
sheetops.cell(1,colx+2).value='细胞涂片诊断'
sheetops.cell(1,colx+1).value='姓名（细胞）'
sheetops.cell(1,colx).value='蜡块诊断部位'
sheetops.cell(1,colx-1).value='蜡块诊断结果'
sheetops.cell(1,colx-2).value='姓名（蜡块）'
workbookops.save(r'C:\Users\Administrator\Desktop\tablet\ops2.xlsx')
workbookops.close() #关闭openpyxl减少内存压力
workbookwax.release_resources()  #xlrd关闭释放内存（release）
#删除不匹配的文件#
workbookops=openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\tablet\ops2.xlsx')
sheetops=workbookops.active
opslen=sheetops.max_row+1
opscol=sheetops.max_column
opslen
opscol
i=2
x=[]
while i<opslen:
    kk=sheetops.cell(i,(opscol-2)).value
    ss=sheetops.cell(i,32).value
    if kk==None or ss==1:
        x.append(i)
    i+=1 
x
xlen=len(x)
i=1
while i<xlen+1:
    sheetops.delete_rows(x[xlen-i])
    i+=1
    i
workbookops.save(r'C:\Users\Administrator\Desktop\tablet\final.xlsx')
workbookops.close()
#添加同一个人的其他数据#
workf=openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\tablet\final.xlsx') 
workwax=openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\tablet\waxblock.xlsx') 
sheetf=workf.active
sheetwax=workwax.active
i=2
rowsf=sheetf.max_row+1
rowswax=sheetwax.max_row+1
dd=[]
ff=[]
zz=[]
while i<rowsf:
    number=sheetf.cell(i,8).value
    ii=1
    c=0
    while ii<rowswax:
        if sheetwax.cell(ii,8).value==sheetf.cell(i,8).value:
            c+=1
            zz.append(ii)
        ii+=1
    if c>1:
        dd.append(i)
        ff.append(c)
    i+=1
    i
maxff=max(ff)
max_col_old=sheetf.max_column
i=1
while i<maxff:
    sheetf.cell(1,max_col_old+i).value='蜡块诊断结果'
    sheetf.cell(1,max_col_old+1+i).value='细胞诊断结果'
    sheetf.cell(1,max_col_old+2+i).value='细胞送检时间'
    i+=1
workbookcell=xlrd.open_workbook(r'C:\Users\Administrator\Desktop\tablet\细胞学2018-2023.xls')  
sheetcell=workbookcell.sheet_by_index(0)    
lenx=len(dd)
k=0
bx=0
celen=sheetcell.nrows
while k<lenx:
    i=1
    while i<ff[k]:
        sheetf.cell(dd[k],max_col_old+i).value=sheetwax.cell(zz[bx],23).value
        ii=0
        while ii<celen:
            if sheetwax.cell(zz[bx-1+i],8).value==sheetcell.cell_value(ii,7) and sheetwax.cell(zz[bx-1+i],14).value==sheetcell.cell_value(ii,8):
                sheetf.cell(dd[k],max_col_old+1+i).value=sheetcell.cell_value(ii,12)
                sheetf.cell(dd[k],max_col_old+2+i).value=sheetcell.cell_value(ii,11)
            ii+=1
        i+=1
    bx+=ff[k]
    k+=1
    k
workf.save(r'C:\Users\Administrator\Desktop\tablet\finalx.xlsx')
workf.close()
workwax.close()
workbookwax.release_resources()